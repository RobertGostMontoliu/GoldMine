from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchWindowException, WebDriverException, ElementClickInterceptedException, NoSuchElementException
import time
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import re
import logging
import argparse
import json
from ui.api_manager import GPMManager, ADSPowerManager, ChromeManager
from ui.components import ConfigManager
from ui.json_path import resource_path

# Configuración del sistema de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Función para cargar la ruta del archivo Excel desde el JSON
def cargar_ruta_excel_json():
    json_path = resource_path("ui//JSON_FILE//UrlRabbyAccounts.json")
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"No se encontró el archivo JSON: {json_path}")
    
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        return data.get("archivo_excel", None)

# Función principal de Rabby Wallets
def rabby_Wallets(password=None, profile_id=None, perfil_id=None, perfil_nombre=None, driver=None, **kwargs):
    """
    Lógica principal del farmeo de Rabby utilizando Selenium.
    """
    
    # Cargar el archivo Excel desde el JSON
    archivo_excel = cargar_ruta_excel_json()
    
    # Validar que el archivo Excel esté definido
    if not archivo_excel:
        raise ValueError("No se proporcionó un archivo Excel para guardar los datos.")
    
    max_retries = 1
    gpm_manager = GPMManager()  # Instancia de GPMManager
    
    # Si perfil_nombre está vacío, obtenlo de GPMManager
    perfil_nombre = gpm_manager.get_profile_name_by_id(profile_id)
    if perfil_nombre is None:
        raise ValueError(f"No se pudo obtener el nombre del perfil para el ID {profile_id}")
    
    for attempt in range(max_retries):
        if driver is None:
            print("No se pudo iniciar el navegador.")
            return
        try:
            if login_rabby(driver, password, perfil_id, perfil_nombre, archivo_excel):
                print("Farmeo exitoso. Cerrando el perfil.")
                break  # Si el farmeo fue exitoso, salimos del bucle
        except Exception as e:
            print(f"Ocurrió un error durante el proceso: {str(e)}")
            if attempt < max_retries - 1:
                print(f"Reintentando... Intento {attempt + 2} de {max_retries}")
            else:
                print("Se alcanzó el número máximo de intentos.")
        finally:
            print("Finalizando el proceso.")
            if profile_id:
                # Cerrar el perfil desde la API
                gpm_manager.close_browser(profile_id)
            
def login_rabby(driver, password, perfil_id, perfil_nombre, archivo_excel):
    
    password = 123456789
    
    try:
        logger.info("Iniciando proceso de login en Rabby Wallet")
        
        # Guardar el handle de la ventana inicial
        main_handle = driver.current_window_handle
        
        # Navegar a la página de bienvenida de Rabby
        driver.get("chrome-extension://acmacodkjbdgmoleebolmdjonilkdbch/index.html#/welcome")
        logger.info("Navegando a la página de bienvenida de Rabby")
        time.sleep(3)
        
        # Esperar y hacer clic en el botón "Next"
        next_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "#root > div > section > footer > button"))
        )
        next_button.click()
        logger.info("Clic en botón 'Next'")
        time.sleep(3)
        
        # Esperar y hacer clic en el botón "Get Started"
        get_started_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "#root > div > section > footer > a > button"))
        )
        get_started_button.click()
        logger.info("Clic en botón 'Get Started'")
        time.sleep(3)
        
        # Esperar y hacer clic en el botón "Create a new seed phrase"
        create_seed_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='root']/div/div[3]/div[1]/div"))
        )
        create_seed_button.click()
        logger.info("Clic en botón 'Create a new seed phrase'")
        time.sleep(3)
        
        # Ingresar contraseña en el textbox
        password_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[@id='password']"))
        )
        password_input.send_keys(password)
        logger.info("Contraseña ingresada")
        time.sleep(3)
        
        # Confirmar contraseña
        confirm_password_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[@id='confirmPassword']"))
        )
        confirm_password_input.send_keys(password)
        logger.info("Contraseña confirmada")
        time.sleep(3)
        
        # Hacer clic en el botón "Next" después de ingresar la contraseña
        next_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "#root > div > div > div > form > div.p-32.pt-0 > button"))
        )
        next_button.click()
        logger.info("Clic en botón 'Next' después de ingresar contraseña")
        
        # Esperar explícitamente 2 segundos
        time.sleep(2)
        
        # Detectar y cambiar a la ventana con la URL específica
        target_url = "chrome-extension://acmacodkjbdgmoleebolmdjonilkdbch/index.html#/mnemonics/create"
        wait_and_switch_to_window(driver, target_url)
        logger.info(f"Cambiado a la ventana con URL: {target_url}")
        
        # Aquí puedes continuar con las interacciones en esta nueva ventana
        # Por ejemplo, mostrar la frase semilla
        try:
            show_seed_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='root']/div/div/div/div[2]/button"))
            )
            show_seed_button.click()
            logger.info("Clic en botón 'Show seed phrase'")
            time.sleep(3)
        except TimeoutException:
            logger.error("No se pudo encontrar el botón 'Show seed phrase'")
            raise
        
        # A partir de aquí, volvemos al flujo normal
        
        # Copiar la frase semilla
        try:
            seed_phrase_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='root']/div/div/div/div[2]/div[2]/div/div"))
            )
            seed_phrase_raw = seed_phrase_element.text
            
            # Procesar la frase semilla para eliminar números y formatear correctamente
            seed_phrase_words = re.findall(r'\d+\.\s*(\w+)', seed_phrase_raw)
            seed_phrase = ' '.join(seed_phrase_words)
            
            logger.info("Frase semilla copiada y procesada exitosamente")
        except TimeoutException:
            logger.error("No se pudo encontrar el elemento de la frase semilla")
            raise
        
        logger.info("Frase semilla copiada, procesada y guardada.")

        # 9. Hacer clic en el botón "I've saved the phrase"
        saved_phrase_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='root']/div/div/div/div[3]/button"))
        )
        saved_phrase_button.click()
        logger.info("Clic en botón 'I've saved the phrase'")
        time.sleep(3)  # Esperamos un poco antes de manejar las ventanas

        # Esperar y cambiar a la ventana específica
        target_url = "chrome-extension://acmacodkjbdgmoleebolmdjonilkdbch/index.html#/import/select-address?hd=HD%20Key%20Tree&keyringId=1"
        wait_and_switch_to_window(driver, target_url)
        logger.info(f"Cambiado a la ventana con URL: {target_url}")
        
        # Esperamos a que la página se cargue completamente
        try:
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='rc-tabs-0-panel-hd']"))
            )
            logger.info("Página cargada completamente")
        except TimeoutException:
            logger.error("No se pudo cargar la página completamente")
        
        # Intentamos encontrar el botón "Switch" con el XPath específico
        try:
            switch_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='rc-tabs-0-panel-hd']/div/div/div/div/div/div[2]/table/tbody/tr[2]/td[1]/button/div"))
            )
            logger.info("Botón 'Switch' detectado")
            switch_button.click()
            logger.info("Clic en botón 'Switch'")
            time.sleep(2)
        except TimeoutException:
            logger.error("No se pudo encontrar el botón 'Switch' con el XPath proporcionado. Intentando continuar...")
        
        # Copiar la dirección de la billetera antes de hacer clic en "DONE"
        try:
            wallet_address_element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='rc-tabs-0-panel-hd']/div/div/div/div/div/div[2]/table/tbody/tr[2]/td[3]/div/span"))
            )
            wallet_address = wallet_address_element.text
            logger.info(f"Dirección de billetera copiada: {wallet_address}")
        except TimeoutException:
            logger.error("No se pudo encontrar la dirección de la billetera")
            wallet_address = "No disponible"

        # Guardar los datos en el Excel
        datos_perfil = (
            perfil_id,  # ID del perfil
            perfil_nombre,   # Nombre del perfil
            wallet_address,  # Dirección de billetera
            seed_phrase,     # Seed phrase obtenida
            "WALLET CREADA CON EXITO"  # Estado
        )
        
        try:
            guardar_datos_wallet(datos_perfil, archivo_excel)
            logger.info("Datos guardados en el archivo Excel")
        except Exception as e:
            logger.error(f"Error al guardar los datos en el archivo Excel: {str(e)}")

        # Hacer clic en el botón final "DONE"
        try:
            done_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='root']/div/div/button"))
            )
            done_button.click()
            logger.info("Clic en botón final 'DONE'")
            logger.info("El navegador se cerrará automáticamente después de hacer clic en 'DONE'")
        except TimeoutException:
            logger.error("No se pudo encontrar el botón 'DONE'")
        except Exception as e:
            logger.error(f"Error inesperado al hacer clic en 'DONE': {str(e)}")

        # No intentamos cerrar el navegador manualmente, ya que se cerrará automáticamente

    except Exception as e:
        logger.error(f"Error durante el proceso de login: {str(e)}")
        datos_perfil = (
            perfil_id,
            perfil_nombre,
            "ERROR",
            "ERROR",
            "ERROR"
        )
        guardar_datos_wallet(datos_perfil, archivo_excel)
        raise

    logger.info("Proceso de creación de wallet completado")
    return True

def wait_for_new_window(driver, current_handle, timeout=30):
    """
    Espera a que aparezca una nueva ventana y devuelve su handle.
    """
    start_time = time.time()
    while time.time() - start_time < timeout:
        handles = driver.window_handles
        new_handles = [h for h in handles if h != current_handle]
        if new_handles:
            return new_handles[0]
        time.sleep(0.5)
    raise TimeoutException("No se abrió una nueva ventana en el tiempo esperado")

def guardar_datos_wallet(datos_perfil, archivo_excel):
    """
    Guarda los datos de un perfil en un archivo Excel.
    :param datos_perfil: Una tupla con los datos en el siguiente orden:
                         (ID, Nombre del perfil, Dirección de billetera, Seed phrase, Estado)
    :param archivo_excel: Ruta del archivo Excel.
    """
    # Verificar si el archivo Excel ya existe
    if not os.path.exists(archivo_excel):
        raise FileNotFoundError(f"El archivo Excel no existe: {archivo_excel}")

    # Cargar el archivo existente
    wb = load_workbook(archivo_excel)
    ws = wb.active

    # Obtener el número de la siguiente fila disponible para usar como ID
    next_row = ws.max_row + 1

    # Actualizar los datos del perfil con el ID correcto
    datos_perfil_actualizados = (next_row - 1,) + datos_perfil[1:]

    # Agregar la tupla como una nueva fila en el Excel
    ws.append(datos_perfil_actualizados)
    
    # Guardar los cambios en el archivo Excel
    wb.save(archivo_excel)
    print(f"Datos guardados en el archivo Excel para el perfil {datos_perfil[1]}")

# Agregar esta nueva función al final del archivo
def wait_and_switch_to_window(driver, target_url, timeout=30):
    """
    Espera y cambia a una ventana específica basada en su URL.
    """
    start_time = time.time()
    while time.time() - start_time < timeout:
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            if driver.current_url == target_url:
                return True
        time.sleep(0.5)
    raise TimeoutException(f"No se encontró la ventana con URL: {target_url}")

# --- MultiThread Integration ---

def start_multithread_farming(profiles, num_concurrent, window_settings):
    """
    Inicia el farmeo de múltiples perfiles usando la infraestructura de multihilo.

    :param profiles: Lista de IDs de perfiles para farmear.
    :param num_concurrent: Número de perfiles concurrentes.
    :param window_settings: Configuraciones de ventana.
    :return: None
    """
    iniciar_farmeo_multiple(profiles, num_concurrent, "rabby", window_settings)

# --- MAIN ---
def parse_arguments():
    parser = argparse.ArgumentParser(description="Ejecutar Rabby Wallet Script")
    parser.add_argument("--browser", type=str, required=True, help="Tipo de navegador (gpm, ads, chrome)")
    parser.add_argument("--profile_id", type=str, help="ID del perfil para GPM")
    parser.add_argument("--profile_name", type=str, help="Nombre del perfil para Chrome")
    parser.add_argument("--win_scale", type=float, help="Escala de la ventana")
    parser.add_argument("--win_pos", type=str, help="Posición de la ventana (x,y)")
    parser.add_argument("--win_size", type=str, help="Tamaño de la ventana (ancho,alto)")
    return parser.parse_args()


if __name__ == "__main__":
    from ui.MultiThreadFarming import iniciar_farmeo_multiple
    from selenium import webdriver

    args = parse_arguments()

    kwargs = {
        "win_scale": args.win_scale,
        "win_pos": args.win_pos,
        "win_size": args.win_size
    }

    profiles = [args.profile_id]  # Aquí deberías pasar una lista de IDs de perfiles si son múltiples
    num_concurrent = 1  # Cambiar a más si deseas farmeo concurrente

    # Crear una instancia del driver de Selenium
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    
    start_multithread_farming(profiles, num_concurrent, kwargs)


