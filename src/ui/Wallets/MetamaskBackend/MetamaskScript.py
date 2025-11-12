from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchWindowException, WebDriverException, ElementClickInterceptedException, NoSuchElementException
import time
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import re
import logging
import argparse
import json
from ui.api_manager import GPMManager, ADSPowerManager, ChromeManager
from ui.components import ConfigManager
from ui.json_path import resource_path

# Configuración del sistema de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Función para cargar la ruta del archivo Excel desde el JSON
def cargar_ruta_excel_json():
    json_path = resource_path("ui//JSON_FILE//UrlMetamaskAccounts.json")
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"No se encontró el archivo JSON: {json_path}")
    
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        return data.get("archivo_excel", None)

# Función principal de Metamask Wallets
def metamask_Wallets(password=None, profile_id=None, perfil_id=None, perfil_nombre=None, driver=None, **kwargs):
    """
    Lógica principal del farmeo de Metamask utilizando Selenium.
    """
    
    # Cargar el archivo Excel desde el JSON
    archivo_excel = cargar_ruta_excel_json()
    
    # Validar que el archivo Excel esté definido
    if not archivo_excel:
        raise ValueError("No se proporcionó un archivo Excel para guardar los datos.")
    
    max_retries = 1
    gpm_manager = GPMManager()  # Instancia de GPMManager
    
    # Si perfil_nombre está vacío, obtenlo de GPMManager
    perfil_nombre = gpm_manager.get_profile_name_by_id(profile_id)
    if perfil_nombre is None:
        raise ValueError(f"No se pudo obtener el nombre del perfil para el ID {profile_id}")
    
    for attempt in range(max_retries):
        if driver is None:
            print("No se pudo iniciar el navegador.")
            return
        try:
            if login_metamask(driver, password, perfil_id, perfil_nombre, archivo_excel):
                print("Farmeo exitoso. Cerrando el perfil.")
                break  # Si el farmeo fue exitoso, salimos del bucle
        except Exception as e:
            print(f"Ocurrió un error durante el proceso: {str(e)}")
            if attempt < max_retries - 1:
                print(f"Reintentando... Intento {attempt + 2} de {max_retries}")
            else:
                print("Se alcanzó el número máximo de intentos.")
        finally:
            print("Finalizando el proceso.")
            if profile_id:
                # Cerrar el perfil desde la API
                gpm_manager.close_browser(profile_id)
            
def login_metamask(driver, password, perfil_id, perfil_nombre, archivo_excel):
    try:
        logger.info("Iniciando proceso de login en Metamask Wallet") 
        
        # Navegar a la página de bienvenida de Metamask
        driver.get("chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html")
        logger.info("Navegando a la página de bienvenida de Metamask")
        time.sleep(3)

        # Intentar interactuar con los elementos de la página
        try:
            # Aceptar términos de Metamask
            terms_checkbox = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='onboarding__terms-checkbox']"))
            )
            driver.execute_script("arguments[0].click();", terms_checkbox)
            time.sleep(1)

            # Crear una nueva billetera
            create_wallet_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='app-content']/div/div[2]/div/div/div/ul/li[2]/button"))
            )
            driver.execute_script("arguments[0].click();", create_wallet_button)
            time.sleep(1)

            # Aceptar términos y condiciones
            agree_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='app-content']/div/div[2]/div/div/div/div[2]/button[2]"))
            )
            agree_button.click()
            time.sleep(1)

            # Ingresar la nueva contraseña
            password_input = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='app-content']/div/div[2]/div/div/div/div[2]/form/div[1]/label/input"))
            )
            driver.execute_script("arguments[0].value = arguments[1];", password_input, password)
            time.sleep(1)

            # Confirmar la contraseña
            confirm_password_input = driver.find_element(By.XPATH, "//*[@id='app-content']/div/div[2]/div/div/div/div[2]/form/div[2]/label/input")
            driver.execute_script("arguments[0].value = arguments[1];", confirm_password_input, password)
            time.sleep(1)

            # Aceptar términos de uso
            terms_checkbox = driver.find_element(By.XPATH, "//*[@id='app-content']/div/div[2]/div/div/div/div[2]/form/div[3]/label/span[1]/input")
            terms_checkbox.click()
            time.sleep(1)

            # Crear la billetera
            create_button = driver.find_element(By.XPATH, "//*[@id='app-content']/div/div[2]/div/div/div/div[2]/form/button")
            create_button.click()
            time.sleep(1)

            # Asegurar la billetera
            secure_wallet_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='app-content']/div/div[2]/div/div/div/div[2]/button[2]"))
            )
            secure_wallet_button.click()
            time.sleep(1)

            # Revelar la frase de recuperación
            reveal_seed_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='app-content']/div/div[2]/div/div/div/div[6]/button"))
            )
            reveal_seed_button.click()
            time.sleep(1)

            # Copiar la seed phrase
            seed_phrase_element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='app-content']/div/div[2]/div/div/div/div[5]/div"))
            )
            seed_phrase = seed_phrase_element.text
            logger.info(f"Seed phrase obtenida: {seed_phrase}")

            # Guardar los datos en el Excel
            datos_perfil = (
                perfil_id,  # ID del perfil
                perfil_nombre,   # Nombre del perfil
                "Dirección de billetera no disponible",  # Dirección de billetera
                seed_phrase,     # Seed phrase obtenida
                "WALLET CREADA CON EXITO"  # Estado
            )
            
            try:
                guardar_datos_wallet(datos_perfil, archivo_excel)
                logger.info("Datos guardados en el archivo Excel")
            except Exception as e:
                logger.error(f"Error al guardar los datos en el archivo Excel: {str(e)}")

            # Hacer clic en "Next"
            next_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='app-content']/div/div[2]/div/div/div/div[6]/div/button"))
            )
            next_button.click()
            time.sleep(1)

            # Ingresar las palabras de la seed phrase solicitadas
            # Aquí necesitarás implementar la lógica para identificar y completar las palabras solicitadas
            # ...

            # Confirmar las palabras ingresadas
            confirm_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='app-content']/div/div[2]/div/div/div/div[5]/button"))
            )
            confirm_button.click()
            time.sleep(1)

            # Completar el proceso de configuración
            got_it_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='app-content']/div/div[2]/div/div/div/div[2]/button"))
            )
            got_it_button.click()
            time.sleep(1)

            # Hacer clic en "Next"
            next_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='app-content']/div/div[2]/div/div/div/div[2]/button"))
            )
            next_button.click()
            time.sleep(1)

            # Hacer clic en "Done"
            done_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='app-content']/div/div[2]/div/div/div/div[2]/button"))
            )
            done_button.click()
            time.sleep(1)

            # Hacer clic en "Got it"
            got_it_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='popover-content']/div/div/section/div[2]/div/div[2]/div/button"))
            )
            got_it_button.click()
            time.sleep(1)

            # Hacer clic en "Receive"
            receive_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='app-content']/div/div[3]/div/div/div/div[1]/div/div[2]/div/button[5]/div"))
            )
            receive_button.click()
            time.sleep(1)

            # Copiar la dirección de la billetera
            wallet_address_element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div[3]/div/section/div/div/p[2]"))
            )
            wallet_address = wallet_address_element.text.replace("\n", "")
            logger.info(f"Dirección de billetera obtenida: {wallet_address}")

            # Actualizar los datos del perfil con la dirección de la billetera
            datos_perfil = (
                perfil_id,
                perfil_nombre,
                wallet_address,
                seed_phrase,
                "WALLET CREADA CON EXITO"
            )
            
            try:
                guardar_datos_wallet(datos_perfil, archivo_excel)
                logger.info("Datos actualizados en el archivo Excel")
            except Exception as e:
                logger.error(f"Error al actualizar los datos en el archivo Excel: {str(e)}")

        except Exception as e:
            logger.error(f"Error durante el proceso de login: {str(e)}")
            datos_perfil = (
                perfil_id,
                perfil_nombre,
                "ERROR",
                "ERROR",
                "ERROR"
            )
            guardar_datos_wallet(datos_perfil, archivo_excel)
            raise

    except Exception as e:
        logger.error(f"Error general: {str(e)}")
        raise

    logger.info("Proceso de creación de wallet completado")
    return True

def wait_for_new_window(driver, current_handle, timeout=30):
    """
    Espera a que aparezca una nueva ventana y devuelve su handle.
    """
    start_time = time.time()
    while time.time() - start_time < timeout:
        handles = driver.window_handles
        new_handles = [h for h in handles if h != current_handle]
        if new_handles:
            return new_handles[0]
        time.sleep(0.5)
    raise TimeoutException("No se abrió una nueva ventana en el tiempo esperado")

def guardar_datos_wallet(datos_perfil, archivo_excel):
    """
    Guarda los datos de un perfil en un archivo Excel.
    :param datos_perfil: Una tupla con los datos en el siguiente orden:
                         (ID, Nombre del perfil, Dirección de billetera, Seed phrase, Estado)
    :param archivo_excel: Ruta del archivo Excel.
    """
    # Verificar si el archivo Excel ya existe
    if not os.path.exists(archivo_excel):
        raise FileNotFoundError(f"El archivo Excel no existe: {archivo_excel}")

    # Cargar el archivo existente
    wb = load_workbook(archivo_excel)
    ws = wb.active

    # Obtener el número de la siguiente fila disponible para usar como ID
    next_row = ws.max_row + 1

    # Actualizar los datos del perfil con el ID correcto
    datos_perfil_actualizados = (next_row - 1,) + datos_perfil[1:]

    # Agregar la tupla como una nueva fila en el Excel
    ws.append(datos_perfil_actualizados)
    
    # Guardar los cambios en el archivo Excel
    wb.save(archivo_excel)
    print(f"Datos guardados en el archivo Excel para el perfil {datos_perfil[1]}")
    
# Agregar esta nueva función al final del archivo
def wait_and_switch_to_window(driver, target_url, timeout=30):
    """
    Espera y cambia a una ventana específica basada en su URL.
    """
    start_time = time.time()
    while time.time() - start_time < timeout:
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            if driver.current_url == target_url:
                return True
        time.sleep(0.5)
    raise TimeoutException(f"No se encontró la ventana con URL: {target_url}")

# --- MultiThread Integration ---

def start_multithread_farming(profiles, num_concurrent, window_settings):
    """
    Inicia el farmeo de múltiples perfiles usando la infraestructura de multihilo.

    :param profiles: Lista de IDs de perfiles para farmear.
    :param num_concurrent: Número de perfiles concurrentes.
    :param window_settings: Configuraciones de ventana.
    :return: None
    """
    iniciar_farmeo_multiple(profiles, num_concurrent, "metamask", window_settings)

# --- MAIN ---
def parse_arguments():
    parser = argparse.ArgumentParser(description="Ejecutar Metamask Wallet Script") 
    parser.add_argument("--browser", type=str, required=True, help="Tipo de navegador (gpm, ads, chrome)")
    parser.add_argument("--profile_id", type=str, help="ID del perfil para GPM")
    parser.add_argument("--profile_name", type=str, help="Nombre del perfil para Chrome")
    parser.add_argument("--win_scale", type=float, help="Escala de la ventana")
    parser.add_argument("--win_pos", type=str, help="Posición de la ventana (x,y)")
    parser.add_argument("--win_size", type=str, help="Tamaño de la ventana (ancho,alto)")
    return parser.parse_args()


if __name__ == "__main__":
    from ui.MultiThreadFarming import iniciar_farmeo_multiple
    from selenium import webdriver

    args = parse_arguments()

    kwargs = {
        "win_scale": args.win_scale,
        "win_pos": args.win_pos,
        "win_size": args.win_size
    }

    profiles = [args.profile_id]  # Aquí deberías pasar una lista de IDs de perfiles si son múltiples
    num_concurrent = 1  # Cambiar a más si deseas farmeo concurrente

    # Crear una instancia del driver de Selenium
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    
    start_multithread_farming(profiles, num_concurrent, kwargs)






