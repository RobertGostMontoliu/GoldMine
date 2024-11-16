from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchWindowException, WebDriverException, ElementClickInterceptedException, NoSuchElementException
import time
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import re
import logging
import argparse
import json
from ui.api_manager import GPMManager, ADSPowerManager, ChromeManager
from ui.components import ConfigManager
from ui.json_path import resource_path

# Configuración del sistema de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Función para cargar la ruta del archivo Excel desde el JSON
def cargar_ruta_excel_json():
    json_path = resource_path("ui\\JSON_FILE\\UrlTonKeeperAccounts.json")
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"No se encontró el archivo JSON: {json_path}")
    
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        return data.get("archivo_excel", None)

# Función principal de TonKeeper Wallets 
def tonkeeper_Wallets(password=None, profile_id=None, perfil_id=None, perfil_nombre=None, driver=None, **kwargs):
    """
    Lógica principal del farmeo de TonKeeper utilizando Selenium.
    """
    
    # Cargar el archivo Excel desde el JSON
    archivo_excel = cargar_ruta_excel_json()
    
    # Validar que el archivo Excel esté definido
    if not archivo_excel:
        raise ValueError("No se proporcionó un archivo Excel para guardar los datos.")
    
    max_retries = 3
    gpm_manager = GPMManager()  # Instancia de GPMManager
    
    # Si perfil_nombre está vacío, obtenlo de GPMManager
    perfil_nombre = gpm_manager.get_profile_name_by_id(profile_id)
    if perfil_nombre is None:
        raise ValueError(f"No se pudo obtener el nombre del perfil para el ID {profile_id}")
    
    for attempt in range(max_retries):
        if driver is None:
            print("No se pudo iniciar el navegador.")
            return
        try:
            # Navegar a la URL inicial
            driver.get("chrome-extension://omaabbefbmiijedngplfjmnooppbclkk/index.html?add_wallet=create-standard")
            time.sleep(1)

            # Paso 2: Click en "Continue"
            continue_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='react-portal-modal-container']/div/div/div/div/div[3]/div/div/div[2]/div/button"))
            )
            continue_button.click()
            time.sleep(1)

            # Paso 3: Guardar las 24 palabras de la seed phrase
            seed_phrase_element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='react-portal-modal-container']/div/div/div/div/div[3]/div/div/div[2]/div[2]"))
            )
            seed_phrase_text = seed_phrase_element.text.split()
            # Filtrar solo las palabras (omitimos los números)
            seed_phrase = [word for word in seed_phrase_text if not word.endswith('.')]
            # Convertir la lista de palabras a una cadena separada por espacios
            seed_phrase_str = ' '.join(seed_phrase)
            time.sleep(1)

            # Paso 4: Click en "Continue"
            continue_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='react-portal-modal-container']/div/div/div/div/div[3]/div/div/div[2]/button"))
            )
            continue_button.click()
            time.sleep(1)

            # Paso 5: Verificar la seed phrase
            for i in range(1, 4):
                label_xpath = f"//*[@id='react-portal-modal-container']/div/div/div/div/div[3]/div/div/div[2]/div[2]/label[{i}]/span"
                input_xpath = f"//*[@id='react-portal-modal-container']/div/div/div/div/div[3]/div/div/div[2]/div[2]/label[{i}]/input"
                
                label_element = WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.XPATH, label_xpath))
                )
                # Extraer el número de la palabra que se necesita
                word_number = int(re.search(r'\d+', label_element.text).group())  # Extraer el número de la palabra
                word_index = word_number - 1  # Convertir a índice de lista
                input_element = driver.find_element(By.XPATH, input_xpath)
                input_element.send_keys(seed_phrase[word_index])
                time.sleep(1)

            # Click en "Continue" después de verificar
            continue_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='react-portal-modal-container']/div/div/div/div/div[3]/div/div/div[2]/div[3]/button"))
            )
            continue_button.click()
            time.sleep(1)

            # Paso 6: Ingresar la contraseña
            password_input = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='react-portal-modal-container']/div[2]/div/div/div/div[3]/div/div/div[2]/form/div[1]/div/input"))
            )
            password_input.send_keys(password)
            time.sleep(1)

            # Paso 7: Reingresar la contraseña
            reenter_password_input = driver.find_element(By.XPATH, "//*[@id='react-portal-modal-container']/div[2]/div/div/div/div[3]/div/div/div[2]/form/div[2]/div/input")
            reenter_password_input.send_keys(password)
            time.sleep(1)

            # Paso 8: Click en "Continue"
            continue_button = driver.find_element(By.XPATH, "//*[@id='react-portal-modal-container']/div[2]/div/div/div/div[3]/div/div/div[2]/form/button")
            continue_button.click()
            time.sleep(1)

            # Paso 9: Click en "Save"
            save_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='react-portal-modal-container']/div/div/div/div/div[3]/div/div/div[2]/form/button"))
            )
            save_button.click()
            time.sleep(2)  # Esperar un poco más para que se complete el guardado

            # Paso 10: Navegar a la página principal de la wallet
            driver.get("chrome-extension://omaabbefbmiijedngplfjmnooppbclkk/index.html")
            time.sleep(2)  # Esperar a que la página cargue

            # Paso 11: Click en el botón "Receive"
            receive_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='body']/div[2]/div[3]/div"))
            )
            receive_button.click()
            time.sleep(1)

            # Paso 12: Copiar la dirección de la wallet
            address_element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='react-portal-modal-container']/div/div[1]/div/div/div[3]/div/div/form/div[2]/div/div/form/div[2]/span"))
            )
            wallet_address = address_element.text

            # Guardar la dirección en el archivo Excel
            guardar_datos_wallet((perfil_id, perfil_nombre, wallet_address, seed_phrase_str, "WALLET CREADA CON EXITO"), archivo_excel)

            logger.info(f"Wallet creada exitosamente. Dirección: {wallet_address}")
            return True

        except Exception as e:
            logger.error(f"Error durante el proceso de creación de wallet: {str(e)}")
            datos_perfil = (
                perfil_id,
                perfil_nombre,
                "ERROR",
                "ERROR",
                "ERROR"
            )
            guardar_datos_wallet(datos_perfil, archivo_excel)
            raise
        finally:
            logger.info("Finalizando el proceso.")
            if profile_id:
                # Cerrar el perfil desde la API
                gpm_manager.close_browser(profile_id)

    logger.info("Proceso de creación de wallet WALLET CREADA CON EXITO")
    return False

# Función para guardar datos en el archivo Excel
def guardar_datos_wallet(datos_perfil, archivo_excel):
    """
    Guarda los datos de un perfil en un archivo Excel.
    :param datos_perfil: Una tupla con los datos en el siguiente orden:
                         (ID, Nombre del perfil, Dirección de billetera, Seed phrase, Estado)
    :param archivo_excel: Ruta del archivo Excel.
    """
    # Verificar si el archivo Excel ya existe
    if not os.path.exists(archivo_excel):
        raise FileNotFoundError(f"El archivo Excel no existe: {archivo_excel}")

    # Cargar el archivo existente
    wb = load_workbook(archivo_excel)
    ws = wb.active

    # Obtener el número de la siguiente fila disponible para usar como ID
    next_row = ws.max_row + 1

    # Actualizar los datos del perfil con el ID correcto
    datos_perfil_actualizados = (next_row - 1,) + datos_perfil[1:]

    # Agregar la tupla como una nueva fila en el Excel
    ws.append(datos_perfil_actualizados)
    
    # Guardar los cambios en el archivo Excel
    wb.save(archivo_excel)
    print(f"Datos guardados en el archivo Excel para el perfil {datos_perfil[1]}")

# Agregar esta nueva función al final del archivo
def wait_and_switch_to_window(driver, target_url, timeout=30):
    """
    Espera y cambia a una ventana específica basada en su URL.
    """
    start_time = time.time()
    while time.time() - start_time < timeout:
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            if driver.current_url == target_url:
                return True
        time.sleep(0.5)
    raise TimeoutException(f"No se encontró la ventana con URL: {target_url}")

# --- MultiThread Integration ---

def start_multithread_farming(profiles, num_concurrent, window_settings):
    """
    Inicia el farmeo de múltiples perfiles usando la infraestructura de multihilo.

    :param profiles: Lista de IDs de perfiles para farmear.
    :param num_concurrent: Número de perfiles concurrentes.
    :param window_settings: Configuraciones de ventana.
    :return: None
    """
    iniciar_farmeo_multiple(profiles, num_concurrent, "TonKeeper", window_settings)

# --- MAIN ---
def parse_arguments():
    parser = argparse.ArgumentParser(description="Ejecutar TonKeeper Wallet Script") 
    parser.add_argument("--browser", type=str, required=True, help="Tipo de navegador (gpm, ads, chrome)")
    parser.add_argument("--profile_id", type=str, help="ID del perfil para GPM")
    parser.add_argument("--profile_name", type=str, help="Nombre del perfil para Chrome")
    parser.add_argument("--win_scale", type=float, help="Escala de la ventana")
    parser.add_argument("--win_pos", type=str, help="Posición de la ventana (x,y)")
    parser.add_argument("--win_size", type=str, help="Tamaño de la ventana (ancho,alto)")
    return parser.parse_args()


if __name__ == "__main__":
    from ui.MultiThreadFarming import iniciar_farmeo_multiple
    from selenium import webdriver

    args = parse_arguments()

    kwargs = {
        "win_scale": args.win_scale,
        "win_pos": args.win_pos,
        "win_size": args.win_size
    }

    profiles = [args.profile_id]  # Aquí deberías pasar una lista de IDs de perfiles si son múltiples
    num_concurrent = 1  # Cambiar a más si deseas farmeo concurrente

    # Crear una instancia del driver de Selenium
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    
    start_multithread_farming(profiles, num_concurrent, kwargs)   

