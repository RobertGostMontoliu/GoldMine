from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchWindowException, WebDriverException, ElementClickInterceptedException, NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import re
import logging
import argparse
import json
from ui.api_manager import GPMManager, ADSPowerManager, ChromeManager
from ui.components import ConfigManager
from ui.json_path import resource_path

# Configuración del sistema de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Función para cargar la ruta del archivo Excel desde el JSON
def cargar_ruta_excel_json():
    json_path = resource_path("ui//JSON_FILE//UrlTonKeeperAccounts.json")
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"No se encontró el archivo JSON: {json_path}")
    
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        return data.get("archivo_excel", None)

# Función principal de Ronin Wallets 
def ronin_Wallets(password=None, profile_id=None, perfil_id=None, perfil_nombre=None, driver=None, **kwargs):
    """
    Lógica principal del farmeo de Ronin Wallet utilizando Selenium.
    """
    
    # Cargar el archivo Excel desde el JSON
    archivo_excel = cargar_ruta_excel_json()
    
    # Validar que el archivo Excel esté definido
    if not archivo_excel:
        raise ValueError("No se proporcionó un archivo Excel para guardar los datos.")
    
    max_retries = 3
    gpm_manager = GPMManager()  # Instancia de GPMManager
    
    # Si perfil_nombre está vacío, obtenlo de GPMManager
    perfil_nombre = gpm_manager.get_profile_name_by_id(profile_id)
    if perfil_nombre is None:
        raise ValueError(f"No se pudo obtener el nombre del perfil para el ID {profile_id}")
    
    for attempt in range(max_retries):
        if driver is None:
            print("No se pudo iniciar el navegador.")
            return
        try:
            # Paso 1: Navegar a la extensión de Ronin Wallet
            logger.info("Navegando a la extensión de Ronin Wallet...")
            driver.get("chrome-extension://fnjhmkhhmkbjkkabndcnnogagogbneec/src/pages/newtab/new-tab.html?#/")
            time.sleep(1)

            # Paso 2: Click en "Create a new wallet"
            logger.info("Haciendo clic en 'Create a new wallet'...")
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app-container"]/div/div/div[2]/div[1]'))).click()
            time.sleep(1)

            # Paso 3: Click en "Continue with recovery phrase"
            logger.info("Haciendo clic en 'Continue with recovery phrase'...")
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app-container"]/div/div/button[2]'))).click()
            time.sleep(1)

            # Paso 4: Ingresar la contraseña
            logger.info("Ingresando contraseña...")
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="app-container"]/div/div/div/form/div/div[1]/div[1]/div/div/input'))).send_keys(password)
            time.sleep(1)

            # Paso 5: Reingresar la contraseña
            logger.info("Reingresando contraseña...")
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="app-container"]/div/div/div/form/div/div[1]/div[2]/div/div/input'))).send_keys(password)
            time.sleep(1)

            # Paso 6: Click en "Continue"
            logger.info("Haciendo clic en 'Continue'...")
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app-container"]/div/div/div/form/div/div[1]/button'))).click()
            time.sleep(1)

            # Paso 7: Click en "Reveal recovery phrase"
            logger.info("Haciendo clic en 'Reveal recovery phrase'...")
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app-container"]/div/div/div/div/div[2]/div[2]/div[2]/button'))).click()
            time.sleep(1)

            # Paso 8: Guardar la seed phrase
            logger.info("Obteniendo seed phrase...")
            seed_phrase_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="app-container"]/div/div/div/div/div[2]/div[2]/div[1]'))
            )
            
            # Obtener el texto y procesarlo para eliminar los números
            seed_text = seed_phrase_element.text
            # Eliminar los números y espacios extra, dejando solo las palabras
            seed_words = [word for word in seed_text.split() if not word[0].isdigit()]
            # Unir las palabras con un solo espacio
            seed_phrase = " ".join(seed_words)
            
            logger.info("Seed phrase obtenida exitosamente")

            # Paso 9: Click en el checkbox usando ActionChains
            logger.info("Haciendo clic en el checkbox...")
            try:
                checkbox = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="app-container"]/div/div/div/div/div[2]/label/span[1]'))
                )
                actions = ActionChains(driver)
                actions.move_to_element(checkbox).click().perform()
                logger.info("Checkbox clickeado exitosamente")
                time.sleep(1)
            except Exception as e:
                logger.error(f"Error al hacer clic en el checkbox: {str(e)}")
                raise

            # Paso 10: Click en "Continue"
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app-container"]/div/div/div/div/div[2]/button[2]'))).click()
            time.sleep(1)

            # Paso 11: Seleccionar las palabras correctas de la seed phrase
            logger.info("Iniciando proceso de confirmación de seed phrase...")
            
            for i in range(1, 4):
                # 1. Obtener el número de palabra solicitada
                word_number_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, f'//*[@id="app-container"]/div/div/div/div[{i}]/div[1]/div[2]'))
                )
                # Extraer el número
                word_number = int(word_number_element.text.replace('#', ''))
                logger.info(f"Se solicita la palabra #{word_number} de la seed phrase")

                # 2. Obtener la palabra correcta de nuestra seed phrase
                correct_word = seed_words[word_number - 1]  # Restamos 1 porque los arrays comienzan en 0
                logger.info(f"La palabra correcta es: {correct_word}")

                # Ejemplo de log para debug
                logger.info(f"Seed words completa: {seed_words}")
                logger.info(f"Buscando palabra en posición {word_number}, que es: {correct_word}")

                # 3. Obtener el contenedor de las opciones
                options_container = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, f'//*[@id="app-container"]/div/div/div/div[{i}]/div[2]'))
                )
                
                # 4. Obtener todos los botones de opciones dentro del contenedor
                option_buttons = options_container.find_elements(By.TAG_NAME, "button")
                
                # 5. Buscar y hacer clic en la opción correcta
                found = False
                for button in option_buttons:
                    if button.text.strip() == correct_word:
                        logger.info(f"Haciendo clic en la opción correcta: {button.text}")
                        button.click()
                        time.sleep(1)
                        found = True
                        break
                
                if not found:
                    logger.error(f"No se encontró la palabra correcta '{correct_word}' entre las opciones disponibles")
                    raise Exception(f"No se pudo encontrar la palabra correcta para la posición {i}")

            logger.info("Proceso de confirmación de seed phrase completado exitosamente")

            # Paso 12: Obtener la dirección de la wallet
            logger.info("Obteniendo dirección de la wallet...")
            wallet_address_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="app-container"]/div/div/div/div/div[1]/div[2]'))
            )
            wallet_address = wallet_address_element.text

            # Guardar datos en Excel
            datos_perfil = (
                perfil_id,
                perfil_nombre,
                wallet_address,
                seed_phrase,
                "WALLET CREADA CON EXITO"
            )
            guardar_datos_wallet(datos_perfil, archivo_excel)

            logger.info(f"Wallet creada exitosamente. Dirección: {wallet_address}")

            # Después de confirmar las 3 palabras
            logger.info("Esperando botón Finish...")
            try:
                finish_button = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="app-container"]/div/div/button'))
                )
                logger.info("Botón Finish detectado. Proceso completado exitosamente")
                
                # No cerramos el navegador aquí, dejamos que GPM lo maneje
                return True
                
            except Exception as e:
                logger.error(f"Error al esperar el botón Finish: {str(e)}")
                raise

        except Exception as e:
            logger.error(f"Error durante el proceso de creación de wallet: {str(e)}")
            datos_perfil = (
                perfil_id,
                perfil_nombre,
                "ERROR",
                seed_phrase,
                "ERROR"
            )
            guardar_datos_wallet(datos_perfil, archivo_excel)
            if attempt < max_retries - 1:
                logger.info(f"Reintentando... Intento {attempt + 2} de {max_retries}")
                continue
            raise
        finally:
            logger.info("Finalizando el proceso.")
            if profile_id:
                try:
                    # Cerrar el perfil desde la API de GPM
                    gpm_manager.close_browser(profile_id)
                except Exception as e:
                    logger.warning(f"Error al cerrar el navegador con GPM: {str(e)}")

    logger.info("Proceso de creación de wallet completado")
    return False

# Función para guardar datos en el archivo Excel
def guardar_datos_wallet(datos_perfil, archivo_excel):
    """
    Guarda los datos de un perfil en un archivo Excel.
    :param datos_perfil: Una tupla con los datos en el siguiente orden:
                         (ID, Nombre del perfil, Dirección de billetera, Seed phrase, Estado)
    :param archivo_excel: Ruta del archivo Excel.
    """
    if not os.path.exists(archivo_excel):
        raise FileNotFoundError(f"El archivo Excel no existe: {archivo_excel}")

    wb = load_workbook(archivo_excel)
    ws = wb.active
    next_row = ws.max_row + 1
    datos_perfil_actualizados = (next_row - 1,) + datos_perfil[1:]
    ws.append(datos_perfil_actualizados)
    wb.save(archivo_excel)
    print(f"Datos guardados en el archivo Excel para el perfil {datos_perfil[1]}")

def wait_and_switch_to_window(driver, target_url, timeout=30):
    """
    Espera y cambia a una ventana específica basada en su URL.
    """
    start_time = time.time()
    while time.time() - start_time < timeout:
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            if driver.current_url == target_url:
                return True
        time.sleep(0.5)
    raise TimeoutException(f"No se encontró la ventana con URL: {target_url}")

# --- MultiThread Integration ---
def start_multithread_farming(profiles, num_concurrent, window_settings):
    """
    Inicia el farmeo de múltiples perfiles usando la infraestructura de multihilo.
    """
    iniciar_farmeo_multiple(profiles, num_concurrent, "Ronin", window_settings)

# --- MAIN ---
def parse_arguments():
    parser = argparse.ArgumentParser(description="Ejecutar Ronin Wallet Script")
    parser.add_argument("--browser", type=str, required=True, help="Tipo de navegador (gpm, ads, chrome)")
    parser.add_argument("--profile_id", type=str, help="ID del perfil para GPM")
    parser.add_argument("--profile_name", type=str, help="Nombre del perfil para Chrome")
    parser.add_argument("--win_scale", type=float, help="Escala de la ventana")
    parser.add_argument("--win_pos", type=str, help="Posición de la ventana (x,y)")
    parser.add_argument("--win_size", type=str, help="Tamaño de la ventana (ancho,alto)")
    return parser.parse_args()

if __name__ == "__main__":
    from ui.MultiThreadFarming import iniciar_farmeo_multiple
    from selenium import webdriver

    args = parse_arguments()

    kwargs = {
        "win_scale": args.win_scale,
        "win_pos": args.win_pos,
        "win_size": args.win_size
    }

    profiles = [args.profile_id]
    num_concurrent = 1

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    
    start_multithread_farming(profiles, num_concurrent, kwargs)