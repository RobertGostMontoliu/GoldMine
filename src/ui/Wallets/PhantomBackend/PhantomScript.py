from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchWindowException, WebDriverException, ElementClickInterceptedException, NoSuchElementException
import time
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import sys
import re
import logging
import argparse
import json
from ui.api_manager import GPMManager, ADSPowerManager, ChromeManager
from ui.components import ConfigManager
from selenium.webdriver.common.action_chains import ActionChains
from ui.json_path import resource_path

# Configuración del sistema de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Función para cargar la ruta del archivo Excel desde el JSON
def cargar_ruta_excel_json():
    json_path = resource_path("ui\\JSON_FILE\\UrlPhantomAccounts.json")
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"No se encontró el archivo JSON: {json_path}")
    
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        return data.get("archivo_excel", None)

# Función principal de Phantom Wallets
def phantom_Wallets(password=None, profile_id=None, perfil_id=None, perfil_nombre=None, driver=None, **kwargs):
    """
    Lógica principal del farmeo de Phantom utilizando Selenium.
    """
    print(f"abrir_y_farmear_blum llamado con profile_id={profile_id}")
          
    # Cargar el archivo Excel desde el JSON
    archivo_excel = cargar_ruta_excel_json()
    
    # Validar que el archivo Excel esté definido
    if not archivo_excel:
        raise ValueError("No se proporcionó un archivo Excel para guardar los datos.")
    
    max_retries = 1
    gpm_manager = GPMManager()  # Instancia de GPMManager
    
    # Si perfil_nombre está vacío, obtenlo de GPMManager
    perfil_nombre = gpm_manager.get_profile_name_by_id(profile_id)
    if perfil_nombre is None:
        raise ValueError(f"No se pudo obtener el nombre del perfil para el ID {profile_id}")
    
    for attempt in range(max_retries):
        if driver is None:
            print("No se pudo iniciar el navegador.")
            return
        try:
            if login_phantom(driver, password, perfil_id, perfil_nombre, archivo_excel):
                print("Farmeo exitoso. Cerrando el perfil.")
                break  # Si el farmeo fue exitoso, salimos del bucle
        except Exception as e:
            print(f"Ocurrió un error durante el proceso: {str(e)}")
            if attempt < max_retries - 1:
                print(f"Reintentando... Intento {attempt + 2} de {max_retries}")
            else:
                print("Se alcanzó el número máximo de intentos.")
        finally:
            print("Finalizando el proceso.")
            if profile_id:
                # Cerrar el perfil desde la API
                gpm_manager.close_browser(profile_id)
            
def login_phantom(driver, password, perfil_id, perfil_nombre, archivo_excel):
    password = 123456789
       
    try:
        # Paso 1: Esperar 2 segundos y navegar a la URL de Phantom
        target_url = "chrome-extension://bfnaelmomeimhlpmgjnjophhpkkoljpa/onboarding.html"
        logger.info("Esperando 2 segundos antes de navegar a Phantom...")
        time.sleep(2)
        
        # Verificar la URL actual
        current_url = driver.current_url
        logger.info(f"URL actual: {current_url}")
        
        if "chrome-extension://acmacodkjbdgmoleebolmdjonilkdbch/offscreen.html" in current_url:
            logger.warning("Detectada ventana en blanco. Navegando a la URL correcta de Phantom...") 
            driver.get(target_url)
        elif target_url not in current_url:
            logger.info("Navegando directamente a la URL de Phantom...") 
            driver.get(target_url)
        else:
            logger.info("Ya estamos en la página correcta de Phantom.") 
        
        # Esperar a que la página de Phantom cargue completamente
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="root"]')))
        
        # Cerrar todas las ventanas excepto la de Phantom
        logger.info("Cerrando todas las ventanas excepto la de Phantom...")
        main_window = driver.current_window_handle
        for handle in driver.window_handles:
            if handle != main_window:
                driver.switch_to.window(handle)
                driver.close()
        driver.switch_to.window(main_window)
        logger.info("Todas las ventanas adicionales han sido cerradas.")
        
        # Verificar si estamos en la página correcta
        if "onboarding.html" not in driver.current_url:
            raise Exception(f"No se pudo acceder a la página de Phantom. URL actual: {driver.current_url}")

        logger.info(f"Página de Phantom cargada correctamente. URL: {driver.current_url}")

        # Paso 2: Click en crear nueva billetera
        logger.info("Haciendo clic en 'Crear nueva billetera'...")
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/main/div[2]/div/div[2]/button[1]'))).click()
        time.sleep(1)
        logger.info("Clic realizado con éxito.")

        # Paso 3: Introducir contraseña
        logger.info("Introduciendo contraseña...")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/main/div[2]/form/div[1]/div[2]/input'))).send_keys(password)
        time.sleep(1)
        logger.info("Contraseña introducida.")

        # Paso 4: Confirmar contraseña
        logger.info("Confirmando contraseña...")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/main/div[2]/form/div[1]/div[2]/div/div/input'))).send_keys(password)
        time.sleep(1)
        logger.info("Contraseña confirmada.")

        # Paso 5: Aceptar términos (hacer clic en el checkbox)
        logger.info("Aceptando términos...")
        checkbox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/main/div[2]/form/div[2]/span')))
        actions = ActionChains(driver)
        actions.move_to_element(checkbox).click().perform()
        time.sleep(1)
        logger.info("Términos aceptados.")

        # Paso 6: Click en continuar
        logger.info("Haciendo clic en 'Continuar'...")
        continue_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/main/div[2]/form/button')))
        actions.move_to_element(continue_button).click().perform()
        time.sleep(1)
        logger.info("Clic en 'Continuar' realizado.")

        # Paso 7: Copiar seed phrase
        logger.info("Copiando seed phrase...")
        seed_phrase_elements = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.XPATH, '//*[@id="root"]/main/div[2]/form/div[1]/div[2]/div/div[2]/div/input'))
        )
        seed_phrase = ' '.join([element.get_attribute('value') for element in seed_phrase_elements])
        time.sleep(1)
        logger.info(f"Seed phrase copiada: {seed_phrase}")

        # Paso 8: Click en checkbox de confirmación
        logger.info("Marcando checkbox de confirmación...")
        confirm_checkbox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/main/div[2]/form/div[2]/span')))
        actions.move_to_element(confirm_checkbox).click().perform()
        time.sleep(1)
        logger.info("Checkbox de confirmación marcado.")

        #cambiando URL
        logger.info("Detectado botón Start, cambiando a URL de login...")
        driver.get("chrome-extension://bfnaelmomeimhlpmgjnjophhpkkoljpa/popup.html")
        time.sleep(2)

        # Hacer clic en recibir
        logger.info("Haciendo clic en recibir...")
        receive_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="home-tab"]/div[1]/div[2]/div/button[1]'))
        )
        receive_button.click()
        time.sleep(1)

        # Hacer clic en Solana
        logger.info("Seleccionando Solana...")
        solana_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="modal"]/div[2]/div/ul/li[1]/button'))
        )
        solana_button.click()
        time.sleep(1)

        # Copiar dirección de wallet
        logger.info("Copiando dirección de wallet...")
        wallet_address_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="modal"]/div[2]/div/div[2]/div[2]/textarea'))
        )
        wallet_address = wallet_address_element.get_attribute('value')
        logger.info(f"Dirección de wallet obtenida: {wallet_address}")

        # Guardar los datos en el Excel
        datos_perfil = (
            perfil_id,
            perfil_nombre,
            wallet_address,  # Ahora usamos la dirección real
            seed_phrase,
            "WALLET CREADA CON ÉXITO"
        )
        
        # TODO: Implement the functionality for "X"
        
        try:
            guardar_datos_wallet(datos_perfil, archivo_excel)
            logger.info("Datos guardados en el archivo Excel")
        except Exception as e:
            logger.error(f"Error al guardar los datos en el archivo Excel: {str(e)}")

    except Exception as e:
        logger.error(f"Error durante el proceso de login: {str(e)}")
        datos_perfil = (
            perfil_id,
            perfil_nombre,
            "ERROR",
            "ERROR",
            "ERROR"
        )
        guardar_datos_wallet(datos_perfil, archivo_excel)
        raise

    logger.info("Proceso de creación de wallet completado")
    return True

def guardar_datos_wallet(datos_perfil, archivo_excel):
    """
    Guarda los datos de un perfil en un archivo Excel.
    :param datos_perfil: Una tupla con los datos en el siguiente orden:
                         (ID, Nombre del perfil, Dirección de billetera, Seed phrase, Estado)
    :param archivo_excel: Ruta del archivo Excel.
    """
    # Verificar si el archivo Excel ya existe
    if not os.path.exists(archivo_excel):
        raise FileNotFoundError(f"El archivo Excel no existe: {archivo_excel}")

    # Cargar el archivo existente
    wb = load_workbook(archivo_excel)
    ws = wb.active

    # Obtener el número de la siguiente fila disponible para usar como ID
    next_row = ws.max_row + 1

    # Actualizar los datos del perfil con el ID correcto
    datos_perfil_actualizados = (next_row - 1,) + datos_perfil[1:]

    # Agregar la tupla como una nueva fila en el Excel
    ws.append(datos_perfil_actualizados)
    
    # Guardar los cambios en el archivo Excel
    wb.save(archivo_excel)
    print(f"Datos guardados en el archivo Excel para el perfil {datos_perfil[1]}")

# Agregar esta nueva función al final del archivo
def wait_and_switch_to_window(driver, target_url, timeout=30):
    """
    Espera y cambia a una ventana específica basada en su URL.
    """
    start_time = time.time()
    while time.time() - start_time < timeout:
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            if driver.current_url == target_url:
                return True
        time.sleep(0.5)
    raise TimeoutException(f"No se encontró la ventana con URL: {target_url}")

# Agregar esta nueva función al final del archivo
def find_window_with_url(driver, target_url):
    """
    Busca una ventana con la URL específica y devuelve su handle si la encuentra.
    """
    for handle in driver.window_handles:
        driver.switch_to.window(handle)
        if driver.current_url == target_url:
            return handle
    return None

# --- MultiThread Integration ---

def start_multithread_farming(profiles, num_concurrent, window_settings):
    """
    Inicia el farmeo de múltiples perfiles usando la infraestructura de multihilo.

    :param profiles: Lista de IDs de perfiles para farmear.
    :param num_concurrent: Número de perfiles concurrentes.
    :param window_settings: Configuraciones de ventana.
    :return: None
    """
    iniciar_farmeo_multiple(profiles, num_concurrent, "phantom", window_settings)

# --- MAIN ---
def parse_arguments():
    parser = argparse.ArgumentParser(description="Ejecutar Phantom Wallet Script") 
    parser.add_argument("--browser", type=str, required=True, help="Tipo de navegador (gpm, ads, chrome)")
    parser.add_argument("--profile_id", type=str, help="ID del perfil para GPM")
    parser.add_argument("--profile_name", type=str, help="Nombre del perfil para Chrome")
    parser.add_argument("--win_scale", type=float, help="Escala de la ventana")
    parser.add_argument("--win_pos", type=str, help="Posición de la ventana (x,y)")
    parser.add_argument("--win_size", type=str, help="Tamaño de la ventana (ancho,alto)")
    return parser.parse_args()


if __name__ == "__main__":
    from ui.MultiThreadFarming import iniciar_farmeo_multiple
    from selenium import webdriver

    args = parse_arguments()

    kwargs = {
        "win_scale": args.win_scale,
        "win_pos": args.win_pos,
        "win_size": args.win_size
    }
    
    profiles = [args.profile_id]  # Aquí deberías pasar una lista de IDs de perfiles si son múltiples
    num_concurrent = 1  # Cambiar a más si deseas farmeo concurrente

    # Crear una instancia del driver de Selenium
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    
    start_multithread_farming(profiles, num_concurrent, kwargs)
