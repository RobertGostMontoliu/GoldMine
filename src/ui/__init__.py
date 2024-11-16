import ui.entrypoint
import requests
import os
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import json
from ui.json_path import resource_path
import sys
import json.decoder
from PyQt5.QtWidgets import QApplication, QMessageBox
from ui.main_window import AdministradorPrincipal
from PyQt5.QtCore import QSettings
from ui.api_manager import GPMManager
from PyQt5.QtGui import QPalette, QColor
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (QMenu, QAction, QWidget, QVBoxLayout, QPushButton, QLabel, QToolButton, QFrame, QHBoxLayout, QTableWidget, QHeaderView, QInputDialog, QMessageBox)
from PyQt5.QtGui import QIcon, QPixmap, QCursor, QFont
from PyQt5.QtCore import Qt, QSize
from ui.config import DEBUG
from ui.appStart import QApplication, load_screen_index, AdministradorPrincipal, set_screen
from ui.key_manager import KeyManager
from datetime import datetime, timedelta
from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel
from PyQt5.QtWidgets import (QMainWindow, QWidget, QHBoxLayout, QVBoxLayout, QStackedWidget, QInputDialog, QMessageBox, QDesktopWidget, QApplication)
from ui.components import create_sidebar, create_profile_button
from ui.pages import (create_home_page, create_telegram_tools_page, create_settings_page, load_profiles_from_api, create_twitter_tools_page)
from ui.AppTheme import apply_theme
from ui.Sonic.SonicPage import SonicPage
from ui.Blum.BlumPage import BlumPage
from ui.FarmeosTelegram.Pawns.PawnsPage import PawnsPage
from ui.FarmeosTelegram.TronKeeper.TronKeeperPage import TronKeeperPage
from ui.Imx.ImxPage import ImxPage
from ui.Wallets.Metamask import MetamaskWindow
from ui.Wallets.Rabby import RabbyWindow
from ui.Wallets.Phantom import PhantomWindow
from ui.Wallets.TonKeeper import TonKeeperWindow
from ui.WolfGame.WolfPage import WolfGamePage
from ui.Wow.WowPage import WowPage
from ui.Wallets.Ronin import RoninWindow
from datetime import datetime
import concurrent.futures
import logging
from matplotlib.pyplot import sca
import pyautogui
from ui.components import ConfigManager
from ui.Sonic.test2 import abrir_y_farmear_claim
from ui.Sonic.Arcadegames import PlayArcadeGames
from ui.twitter.TwitterBackend.TwitterLogin import interactuar_con_twitter
import threading
from PyQt5.QtCore import QThread
from ui.Blum.BlumBackend.BlumClaim import abrir_y_farmear_blum
from ui.FarmeosTelegram.Pawns.PawnsBackend.PawnsScript import abrir_y_farmear_pawns
from ui.FarmeosTelegram.TronKeeper.TronBackend.TronScript import abrir_y_farmear_tronkeeper
from ui.Wallets.RabbyBackend.RabbyScript import rabby_Wallets
from ui.Wallets.MetamaskBackend.MetamaskScript import metamask_Wallets
from ui.Wallets.TonKeeperBackend.TonKeeperScript import tonkeeper_Wallets
from ui.Wallets.PhantomBackend.PhantomScript import phantom_Wallets
from ui.WolfGame.Backend.WolfGameScript import WolfGame
from ui.Wow.Backend.WowScript import Wow
from ui.Wallets.RoninBackend.RoninScript import ronin_Wallets
from ui.TelegramTool.telegram_backend.TelegramSignIn import telegramTool
from ui.TelegramTool.telegram_backend.TelegramPassword import telegramToolPassword
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QLineEdit, QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog, QComboBox, QDialog, QApplication, QAbstractItemView)
from PyQt5.QtCore import Qt, QUrl
from PyQt5.QtGui import QFont, QPixmap
from PyQt5.QtWebEngineWidgets import QWebEngineView
from ui.TelegramTool.TelegramTool import TelegramToolPage
from ui.twitter.TwitterPage import TwitterToolPage
from ui.api_manager import GPMManager, ADSPowerManager, ChromeManager
from ui.style_text import apply_text_input_style
from ui.style_box import apply_button_style
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QTableWidget, QTableWidgetItem, QHeaderView, QLineEdit, QComboBox, QFrame, QGroupBox, QPushButton, QSpacerItem, QSizePolicy, QAbstractItemView, QCheckBox)
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QLineEdit, QTextEdit
from PyQt5.QtWidgets import QLineEdit
from PyQt5.QtWidgets import (QPushButton, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QGroupBox, QCheckBox, QAbstractItemView, QMessageBox,QTableWidgetItem, QTextEdit, QDialog)
from PyQt5.QtGui import QColor
from PyQt5.QtCore import QTimer
from ui.PlantillaAirdrop import PlantillaAirdrop
from ui.LogWindow import LogWindow
from ui.MultiThreadFarming import iniciar_farmeo_multiple
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchWindowException, WebDriverException, ElementClickInterceptedException, NoSuchElementException
import argparse
import cv2
import numpy as np
import math
from PIL import Image
import io
from PyQt5.QtWidgets import (QPushButton, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QGroupBox, QCheckBox, QAbstractItemView, QMessageBox,QTableWidgetItem)
from ui.Log_JSON import Logger
from PyQt5.QtWidgets import QTableWidgetItem
from ui.Sonic.RabbyHandler import RabbyHandler
from selenium.common.exceptions import TimeoutException, NoSuchWindowException, StaleElementReferenceException, WebDriverException
from PyQt5.QtWidgets import (QPushButton, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QMessageBox, QGroupBox, QCheckBox, QAbstractItemView, QMessageBox,QTableWidgetItem, QDialog)
from selenium.common.exceptions import TimeoutException, NoSuchWindowException, WebDriverException, ElementClickInterceptedException
from selenium.webdriver.common.action_chains import ActionChains
from PyQt5.QtWidgets import QMessageBox
import pytest
from PyQt5.QtWidgets import (QTextEdit, QPushButton, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QMessageBox, QGroupBox, QAbstractItemView, QComboBox, QWidget, QDialog, QTableWidgetItem)
import re
from ui.style_box import apply_text_input_style
from ui.style_box import apply_button_style, apply_button_grey_style
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, JavascriptException
import random
import traceback
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException, StaleElementReferenceException, ElementNotInteractableException
from PyQt5.QtWidgets import (QTextEdit, QPushButton, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QMessageBox, QGroupBox, QAbstractItemView, QTableWidgetItem, QDialog)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QPixmap
from selenium.common.exceptions import TimeoutException
from PyQt5.QtWidgets import QVBoxLayout, QFileDialog, QAbstractItemView, QWidget, QMessageBox, QHBoxLayout, QLabel, QLineEdit, QCheckBox, QPushButton, QGroupBox, QTableWidgetItem, QDialog
from openpyxl import Workbook
from PyQt5.QtWidgets import QVBoxLayout, QFileDialog, QAbstractItemView, QWidget, QMessageBox, QHBoxLayout, QLabel, QLineEdit, QCheckBox, QPushButton, QGroupBox, QDialog, QTableWidgetItem
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl import load_workbook
from PyQt5.QtWidgets import (QPushButton, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QMessageBox, QGroupBox, QCheckBox, QAbstractItemView)
from ui.WolfGame.Backend import WolfExcel as excel
from collections import deque
import os.path
from ui.WolfGame.Backend import WolfAuto
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
from ui.PaymentPage.PayPage import PayPage
